"""
Модуль для выгрузки данных КРД в Excel
✅ СТРУКТУРИРОВАННЫЙ ВЫВОД: Каждая КРД занимает несколько строк (основные данные + связанные записи)
✅ ФИЛЬТРАЦИЯ: Выгружаются только те столбцы, которые выбраны в шаблоне отчета
✅ ЧИСТЫЕ ЗАГОЛОВКИ: Убраны префиксы названий таблиц из заголовков Excel
✅ БЕЗОПАСНОСТЬ: Все SQL-запросы используют bindValue
✅ ФОРМАТИРОВАНИЕ: Корректные aRGB цвета для openpyxl
"""
from PyQt6.QtSql import QSqlQuery
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
import os
import traceback


class KrdExcelExporter:
    """Экспорт данных КРД в Excel с поддержкой конфигурации отчета"""

    # Описание всех доступных полей и секций
    AVAILABLE_FIELDS = {
        "social_data": {
            "title": "Социально-демографические данные",
            "fields": [
                ("krd_number", "№ КРД"), ("tab_number", "Табельный номер"), ("personal_number", "Личный номер"),
                ("category_name", "Категория военнослужащего"), ("rank_name", "Воинское звание"),
                ("surname", "Фамилия"), ("name", "Имя"), ("patronymic", "Отчество"),
                ("birth_date", "Дата рождения"), ("birth_place_town", "Населенный пункт места рождения"),
                ("birth_place_district", "Административный район места рождения"), ("birth_place_region", "Субъект (регион) места рождения"),
                ("birth_place_country", "Страна места рождения"), ("drafted_by_commissariat", "Наименование комиссариата"),
                ("draft_date", "Дата призыва"), ("povsk", "Наименование ПОВСК"), ("selection_date", "Дата отбора"),
                ("education", "Образование"), ("criminal_record", "Сведения о судимости"),
                ("social_media_account", "Аккаунт в социальных сетях"), ("bank_card_number", "Номер банковской карты"),
                ("passport_series", "Серия паспорта"), ("passport_number", "Номер паспорта"),
                ("passport_issue_date", "Дата выдачи паспорта"), ("passport_issued_by", "Кем выдан паспорт"),
                ("military_id_series", "Серия военного билета"), ("military_id_number", "Номер военного билета"),
                ("military_id_issue_date", "Дата выдачи военного билета"), ("military_id_issued_by", "Кем выдан военный билет"),
                ("appearance_features", "Особенности внешности"), ("personal_marks", "Личные приметы"),
                ("military_contacts", "Контакты в/с"), ("relatives_info", "Сведения о близких родственниках"),
            ]
        },
        "addresses": {
            "title": "Адреса проживания",
            "fields": [
                ("region", "Субъект РФ"), ("district", "Административный район"), ("town", "Населенный пункт"),
                ("street", "Улица"), ("house", "Дом"), ("building", "Корпус"), ("letter", "Литер"),
                ("apartment", "Квартира"), ("room", "Комната"), ("check_date", "Дата адресной проверки"),
                ("check_result", "Результат адресной проверки")
            ]
        },
        "incoming_orders": {
            "title": "Входящие поручения на розыск",
            "fields": [
                ("initiator_full_name", "Инициатор розыска"), ("order_date", "Исходящая дата поручения"),
                ("order_number", "Исходящий номер поручения"), ("receipt_date", "Дата поступления в ВК"),
                ("receipt_number", "Входящий номер в ВК"), ("postal_index", "Индекс"),
                ("postal_region", "Субъект РФ"), ("postal_district", "Административный район"),
                ("postal_town", "Населенный пункт"), ("postal_street", "Улица"), ("postal_house", "Дом"),
                ("initiator_contacts", "Контакты источника"), ("our_response_date", "Дата ответа ВК"),
                ("our_response_number", "Исходящий номер ответа ВК"), ("military_unit_name", "Военное управление инициатора")
            ]
        },
        "service_places": {
            "title": "Места службы",
            "fields": [
                ("place_name", "Наименование места службы"), ("military_unit_name", "Военное управление места службы"),
                ("garrison_name", "Гарнизон места службы"), ("position_name", "Воинская должность"),
                ("commanders", "Командиры (начальники)"), ("postal_index", "Индекс"),
                ("postal_region", "Субъект РФ"), ("postal_town", "Населенный пункт"), ("postal_street", "Улица"),
                ("postal_house", "Дом"), ("place_contacts", "Контакты места службы")
            ]
        },
        "soch_episodes": {
            "title": "Сведения о СОЧ",
            "fields": [
                ("soch_date", "Дата СОЧ"), ("soch_location", "Место СОЧ"),
                ("order_date_number", "Дата и номер приказа о СОЧ"), ("witnesses", "Очевидцы СОЧ"),
                ("reasons", "Вероятные причины СОЧ"), ("weapon_info", "Сведения о наличии оружия"),
                ("clothing", "Во что был одет"), ("movement_options", "Варианты движения"),
                ("search_date", "Дата розыска"), ("found_by", "Кем разыскан"),
                ("notification_date", "Дата уведомления"), ("notification_number", "Номер уведомления")
            ]
        },
        "outgoing_requests": {
            "title": "Исходящие запросы и поручения",
            "fields": [
                ("request_type_name", "Наименование запроса"), ("recipient_name", "Наименование адресата"),
                ("military_unit_name", "Военное управление адресата"), ("issue_date", "Исходящая дата"),
                ("issue_number", "Исходящий номер"), ("postal_index", "Индекс"),
                ("postal_region", "Субъект РФ"), ("postal_town", "Населенный пункт"), ("postal_street", "Улица"),
                ("postal_house", "Дом"), ("recipient_contacts", "Контакты")
            ]
        }
    }

    def __init__(self, db_connection, krd_id=None, report_config=None):
        self.db = db_connection
        self.krd_id = krd_id
        self.wb = Workbook()
        self.report_config = report_config or self._get_default_config()
        
        # ✅ Стили: формат aRGB (8 символов, без решетки #)
        self.header_font = Font(bold=True, size=11, color="FFFFFFFF")
        self.header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
        self.section_font = Font(bold=True, size=12, color="FF000000")
        self.section_fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
        self.krd_header_font = Font(bold=True, size=13, color="FF003366")
        self.krd_header_fill = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")

        self.cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        self.thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"), 
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        self.temp_image_files = []

    def _get_default_config(self):
        return {
            "sections": ["social_data", "addresses", "incoming_orders", "service_places", "soch_episodes", "outgoing_requests"],
            "fields": {}
        }

    def _cleanup_temp_files(self):
        for temp_file in self.temp_image_files:
            try:
                if os.path.exists(temp_file): os.remove(temp_file)
            except Exception: pass
        self.temp_image_files = []

    def _format_date(self, date_value):
        if date_value:
            try:
                return date_value.toString("dd.MM.yyyy") if hasattr(date_value, 'toString') else str(date_value)
            except Exception:
                return str(date_value)
        return ""

    def _log(self, level, message):
        """Вспомогательный метод для вывода в консоль"""
        print(f"[{level}] {message}")

    def export_multiple_krd_to_excel(self, file_path, krd_ids=None):
        """✅ ЭКСПОРТ СПИСКА: Все данные на одном листе в виде плоской таблицы"""
        try:
            self._log("EXPORT", "🚀 " + "="*60)
            self._log("EXPORT", "🚀 НАЧАЛО ЭКСПОРТА СПИСКА КРД")
            self._log("EXPORT", "🚀 " + "="*60)
            
            if krd_ids is None:
                krd_ids = self.report_config.get("krd_ids", [])
            if not krd_ids:
                raise Exception("Не указан список КРД для экспорта")

            self._log("EXPORT", f"📋 Всего записей для экспорта: {len(krd_ids)}")
            self.wb.remove(self.wb.active)
            ws = self.wb.create_sheet("Список КРД")
            
            # Вызываем метод для заполнения таблицы
            current_row = self._fill_flat_list_table(ws, krd_ids)
            
            self._adjust_column_widths(ws)
            
            self._log("EXPORT", f"💾 Сохранение файла: {file_path}")
            self.wb.save(file_path)
            self._cleanup_temp_files()
            
            self._log("EXPORT", "✅ " + "="*60)
            self._log("EXPORT", "✅ ЭКСПОРТ ЗАВЕРШЁН УСПЕШНО")
            self._log("EXPORT", "✅ " + "="*60)
            return True
        except Exception as e:
            self._log("ERROR", f"✗ ОШИБКА ЭКСПОРТА: {e}")
            traceback.print_exc()
            self._cleanup_temp_files()
            raise

    def _fill_flat_list_table(self, ws, krd_ids):
        """
        Заполнение Excel-листа в виде ЕДИНОГО СПИСКА.
        ✅ Убраны префиксы таблиц из заголовков.
        ✅ Корректная работа с множественными данными (адреса, места службы).
        """
        row = 1
        sections = self.report_config.get("sections", [])
        fields_config = self.report_config.get("fields", {})

        self._log("DEBUG", "📐 Формирование заголовков таблицы...")
        
        # Формируем единый список колонок для шапки
        columns = []
        
        # 1. Основные данные (Social Data)
        all_soc = self.AVAILABLE_FIELDS["social_data"]["fields"]
        
        # Проверяем, есть ли настройки полей для этой секции в конфиге
        if "social_data" in fields_config:
            soc_fields = fields_config["social_data"]
            selected_soc = [(k, v) for k, v in all_soc if k in soc_fields]
        else:
            selected_soc = all_soc
            
        columns.extend([(k, v, "social") for k, v in selected_soc])

        # 2. Секции с множественными данными
        for sec_key in ["addresses", "service_places", "incoming_orders", "soch_episodes", "outgoing_requests"]:
            if sec_key in sections:
                all_sec = self.AVAILABLE_FIELDS[sec_key]["fields"]
                
                # Проверяем настройки полей для текущей секции
                if sec_key in fields_config:
                    sec_fields = fields_config[sec_key]
                    selected_sec = [(k, v) for k, v in all_sec if k in sec_fields]
                else:
                    selected_sec = all_sec
                
                # ✅ ИСПРАВЛЕНО: Убран префикс с названием таблицы (было f"[{sec_key.upper()}] {v}")
                # Теперь добавляем просто (k, v, sec_key)
                columns.extend([(k, v, sec_key) for k, v in selected_sec])

        self._log("DEBUG", f"📊 Сформировано {len(columns)} колонок")

        # Рисуем шапку
        for c, (_, field_name, _) in enumerate(columns, 1):
            cell = ws.cell(row=row, column=c, value=field_name)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
        row += 1

        # Заполняем данными
        for krd_idx, krd_id in enumerate(krd_ids, 1):
            self._log("DATA", f"📝 Обработка КРД-{krd_id} ({krd_idx}/{len(krd_ids)})...")
            
            # Загружаем данные
            social = self._load_social_data_for_krd(krd_id)
            if not social:
                self._log("WARN", f"⚠️ Нет социальных данных для КРД-{krd_id}. Пропуск.")
                continue

            # Загружаем списки связанных данных (если секция включена в конфиг)
            addresses = self._load_addresses_for_krd(krd_id) if "addresses" in sections else []
            service_places = self._load_service_places_for_krd(krd_id) if "service_places" in sections else []
            incoming_orders = self._load_incoming_orders_for_krd(krd_id) if "incoming_orders" in sections else []
            soch_episodes = self._load_soch_episodes_for_krd(krd_id) if "soch_episodes" in sections else []
            outgoing_requests = self._load_outgoing_requests_for_krd(krd_id) if "outgoing_requests" in sections else []

            # Определяем, сколько строк нужно для этой КРД (максимум из всех связанных списков)
            list_counts = [len(lst) for lst in [addresses, service_places, incoming_orders, soch_episodes, outgoing_requests]]
            max_rows = max([1] + list_counts)
            
            self._log("DEBUG", f"   📏 Строк для КРД-{krd_id}: {max_rows}")

            # Заполняем строки
            for r_idx in range(max_rows):
                for c_idx, (key, _, sec_type) in enumerate(columns, 1):
                    value = ""
                    
                    # Логика заполнения в зависимости от типа колонки
                    if sec_type == "social":
                        # Соц. данные только в самой первой строке (r_idx == 0)
                        if r_idx == 0:
                            val = social.get(key)
                            value = self._format_date(val) if key.endswith('_date') else (val or '')
                        else:
                            value = "" # Остальные строки пустые для соц. данных
                    
                    elif sec_type == "addresses" and len(addresses) > 0:
                        if r_idx < len(addresses):
                            val = addresses[r_idx].get(key)
                            value = self._format_date(val) if key.endswith('_date') else (val or '')
                    
                    elif sec_type == "service_places" and len(service_places) > 0:
                        if r_idx < len(service_places):
                            val = service_places[r_idx].get(key)
                            value = val or '' 
                    
                    elif sec_type == "incoming_orders" and len(incoming_orders) > 0:
                        if r_idx < len(incoming_orders):
                            val = incoming_orders[r_idx].get(key)
                            value = self._format_date(val) if key.endswith('_date') else (val or '')
                            
                    elif sec_type == "soch_episodes" and len(soch_episodes) > 0:
                        if r_idx < len(soch_episodes):
                            val = soch_episodes[r_idx].get(key)
                            value = self._format_date(val) if key.endswith('_date') else (val or '')

                    elif sec_type == "outgoing_requests" and len(outgoing_requests) > 0:
                        if r_idx < len(outgoing_requests):
                            val = outgoing_requests[r_idx].get(key)
                            value = self._format_date(val) if key.endswith('_date') else (val or '')

                    # Записываем в ячейку
                    ws.cell(row=row, column=c_idx, value=value).border = self.thin_border
                
                row += 1
            
            # Пустая строка-разделитель между КРД
            row += 1

        self._log("DEBUG", f"🏁 Заполнение таблицы завершено. Последняя строка: {row}")
        return row

    # ================= ЗАГРУЗЧИКИ ДАННЫХ (БЕЗОПАСНЫЕ SQL) =================
    def _load_social_data_for_krd(self, krd_id):
        q = QSqlQuery(self.db)
        q.prepare("""SELECT kr.id as krd_id, s.tab_number, s.personal_number, c.name as category_name, r.name as rank_name,
            s.surname, s.name, s.patronymic, s.birth_date, s.birth_place_town, s.birth_place_district, s.birth_place_region, s.birth_place_country,
            s.drafted_by_commissariat, s.draft_date, s.povsk, s.selection_date, s.education, s.criminal_record, s.social_media_account, s.bank_card_number,
            s.passport_series, s.passport_number, s.passport_issue_date, s.passport_issued_by, s.military_id_series, s.military_id_number,
            s.military_id_issue_date, s.military_id_issued_by, s.appearance_features, s.personal_marks, s.federal_search_info, s.military_contacts, s.relatives_info
            FROM krd.social_data s 
            LEFT JOIN krd.krd kr ON s.krd_id = kr.id 
            LEFT JOIN krd.categories c ON s.category_id = c.id 
            LEFT JOIN krd.ranks r ON s.rank_id = r.id
            WHERE s.krd_id = :krd_id ORDER BY s.id DESC LIMIT 1""")
        q.bindValue(":krd_id", krd_id)
        if not q.exec():
            self._log("SQL", f"❌ Ошибка SQL (social_data): {q.lastError().text()}")
            return {}
        if q.next():
            return {
                "krd_number": f"КРД-{q.value('krd_id')}", "tab_number": q.value("tab_number") or "", "personal_number": q.value("personal_number") or "",
                "category_name": q.value("category_name") or "", "rank_name": q.value("rank_name") or "", "surname": q.value("surname") or "",
                "name": q.value("name") or "", "patronymic": q.value("patronymic") or "", "birth_date": q.value("birth_date"),
                "birth_place_town": q.value("birth_place_town") or "", "birth_place_district": q.value("birth_place_district") or "",
                "birth_place_region": q.value("birth_place_region") or "", "birth_place_country": q.value("birth_place_country") or "",
                "drafted_by_commissariat": q.value("drafted_by_commissariat") or "", "draft_date": q.value("draft_date"),
                "povsk": q.value("povsk") or "", "selection_date": q.value("selection_date"), "education": q.value("education") or "",
                "criminal_record": q.value("criminal_record") or "", "social_media_account": q.value("social_media_account") or "",
                "bank_card_number": q.value("bank_card_number") or "", "passport_series": q.value("passport_series") or "",
                "passport_number": q.value("passport_number") or "", "passport_issue_date": q.value("passport_issue_date"),
                "passport_issued_by": q.value("passport_issued_by") or "", "military_id_series": q.value("military_id_series") or "",
                "military_id_number": q.value("military_id_number") or "", "military_id_issue_date": q.value("military_id_issue_date"),
                "military_id_issued_by": q.value("military_id_issued_by") or "", "appearance_features": q.value("appearance_features") or "",
                "personal_marks": q.value("personal_marks") or "", "federal_search_info": q.value("federal_search_info") or "",
                "military_contacts": q.value("military_contacts") or "", "relatives_info": q.value("relatives_info") or ""
            }
        return {}

    def _load_addresses_for_krd(self, krd_id):
        q = QSqlQuery(self.db)
        q.prepare("SELECT * FROM krd.addresses WHERE krd_id = :krd_id ORDER BY id DESC")
        q.bindValue(":krd_id", krd_id)
        if not q.exec():
            self._log("SQL", f"❌ Ошибка SQL (addresses): {q.lastError().text()}")
            return []
        results = []
        record = q.record()
        field_count = record.count()
        field_names = [record.fieldName(i) for i in range(field_count)]
        
        while q.next():
            row_dict = {}
            for field_name in field_names:
                row_dict[field_name] = q.value(field_name) or ""
            results.append(row_dict)
            
        self._log("DATA", f"   📍 Загружено адресов: {len(results)}")
        return results

    def _load_service_places_for_krd(self, krd_id):
        q = QSqlQuery(self.db)
        q.prepare("""SELECT s.place_name, m.name as military_unit_name, g.name as garrison_name, p.name as position_name, s.commanders, 
            s.postal_index, s.postal_region, s.postal_district, s.postal_town, s.postal_street, s.postal_house, s.place_contacts 
            FROM krd.service_places s 
            LEFT JOIN krd.military_units m ON s.military_unit_id = m.id 
            LEFT JOIN krd.garrisons g ON s.garrison_id = g.id 
            LEFT JOIN krd.positions p ON s.position_id = p.id 
            WHERE s.krd_id = :krd_id ORDER BY s.id DESC""")
        q.bindValue(":krd_id", krd_id)
        if not q.exec():
            self._log("SQL", f"❌ Ошибка SQL (service_places): {q.lastError().text()}")
            return []
        results = []
        while q.next():
            results.append({
                "place_name": q.value("place_name") or "", "military_unit_name": q.value("military_unit_name") or "",
                "garrison_name": q.value("garrison_name") or "", "position_name": q.value("position_name") or "",
                "commanders": q.value("commanders") or "", "postal_index": q.value("postal_index") or "",
                "postal_region": q.value("postal_region") or "", "postal_district": q.value("postal_district") or "",
                "postal_town": q.value("postal_town") or "", "postal_street": q.value("postal_street") or "",
                "postal_house": q.value("postal_house") or "", "place_contacts": q.value("place_contacts") or ""
            })
        self._log("DATA", f"   🎖️ Загружено мест службы: {len(results)}")
        return results

    def _load_incoming_orders_for_krd(self, krd_id):
        q = QSqlQuery(self.db)
        q.prepare("""SELECT i.initiator_full_name, i.order_date, i.order_number, i.receipt_date, i.receipt_number, 
            i.postal_index, i.postal_region, i.postal_district, i.postal_town, i.postal_street, i.postal_house, 
            i.initiator_contacts, i.our_response_date, i.our_response_number, m.name as military_unit_name 
            FROM krd.incoming_orders i 
            LEFT JOIN krd.military_units m ON i.military_unit_id = m.id 
            WHERE i.krd_id = :krd_id ORDER BY i.receipt_date DESC""")
        q.bindValue(":krd_id", krd_id)
        if not q.exec():
            self._log("SQL", f"❌ Ошибка SQL (incoming_orders): {q.lastError().text()}")
            return []
        results = []
        while q.next():
            results.append({
                "initiator_full_name": q.value("initiator_full_name") or "", "order_date": q.value("order_date"),
                "order_number": q.value("order_number") or "", "receipt_date": q.value("receipt_date"),
                "receipt_number": q.value("receipt_number") or "", "postal_index": q.value("postal_index") or "",
                "postal_region": q.value("postal_region") or "", "postal_district": q.value("postal_district") or "",
                "postal_town": q.value("postal_town") or "", "postal_street": q.value("postal_street") or "",
                "postal_house": q.value("postal_house") or "", "initiator_contacts": q.value("initiator_contacts") or "",
                "our_response_date": q.value("our_response_date"), "our_response_number": q.value("our_response_number") or "",
                "military_unit_name": q.value("military_unit_name") or ""
            })
        self._log("DATA", f"   📬 Загружено поручений: {len(results)}")
        return results

    def _load_soch_episodes_for_krd(self, krd_id):
        q = QSqlQuery(self.db)
        q.prepare("""SELECT soch_date, soch_location, order_date_number, witnesses, reasons, weapon_info, clothing, 
            movement_options, search_date, found_by, notification_date, notification_number 
            FROM krd.soch_episodes 
            WHERE krd_id = :krd_id ORDER BY soch_date DESC""")
        q.bindValue(":krd_id", krd_id)
        if not q.exec():
            self._log("SQL", f"❌ Ошибка SQL (soch_episodes): {q.lastError().text()}")
            return []
        results = []
        while q.next():
            results.append({
                "soch_date": q.value("soch_date"), "soch_location": q.value("soch_location") or "",
                "order_date_number": q.value("order_date_number") or "", "witnesses": q.value("witnesses") or "",
                "reasons": q.value("reasons") or "", "weapon_info": q.value("weapon_info") or "",
                "clothing": q.value("clothing") or "", "movement_options": q.value("movement_options") or "",
                "search_date": q.value("search_date"), "found_by": q.value("found_by") or "",
                "notification_date": q.value("notification_date"), "notification_number": q.value("notification_number") or ""
            })
        self._log("DATA", f"   ⚠️ Загружено эпизодов СОЧ: {len(results)}")
        return results

    def _load_outgoing_requests_for_krd(self, krd_id):
        q = QSqlQuery(self.db)
        q.prepare("""SELECT r.recipient_name, r.issue_date, r.issue_number, r.postal_index, r.postal_region, 
            r.postal_district, r.postal_town, r.postal_street, r.postal_house, r.recipient_contacts, 
            t.name as request_type_name, m.name as military_unit_name 
            FROM krd.outgoing_requests r 
            LEFT JOIN krd.request_types t ON r.request_type_id = t.id 
            LEFT JOIN krd.military_units m ON r.military_unit_id = m.id 
            WHERE r.krd_id = :krd_id ORDER BY r.issue_date DESC""")
        q.bindValue(":krd_id", krd_id)
        if not q.exec():
            self._log("SQL", f"❌ Ошибка SQL (outgoing_requests): {q.lastError().text()}")
            return []
        results = []
        while q.next():
            results.append({
                "request_type_name": q.value("request_type_name") or "", "recipient_name": q.value("recipient_name") or "",
                "military_unit_name": q.value("military_unit_name") or "", "issue_date": q.value("issue_date"),
                "issue_number": q.value("issue_number") or "", "postal_index": q.value("postal_index") or "",
                "postal_region": q.value("postal_region") or "", "postal_town": q.value("postal_town") or "",
                "postal_street": q.value("postal_street") or "", "postal_house": q.value("postal_house") or "",
                "recipient_contacts": q.value("recipient_contacts") or ""
            })
        self._log("DATA", f"   📤 Загружено запросов: {len(results)}")
        return results

    def _adjust_column_widths(self, ws):
        """Автоматическая подстройка ширины колонок"""
        self._log("FORMAT", "📏 Автоподстройка ширины колонок...")
        for col_num in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_num)
            for row_num in range(1, min(ws.max_row + 1, 500)):
                cell = ws.cell(row=row_num, column=col_num)
                if cell.value and not isinstance(cell, MergedCell):
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 3, 40)
            ws.column_dimensions[column_letter].width = adjusted_width
        self._log("FORMAT", "✅ Ширина колонок настроена")

    # ================= СТАРЫЕ МЕТОДЫ (ОБРАТНАЯ СОВМЕСТИМОСТЬ) =================
    def export_to_excel(self, file_path):
        """Экспорт одной КРД (устаревший метод)"""
        self._log("EXPORT", f"📊 НАЧАЛО ЭКСПОРТА ОДНОЙ КРД-{self.krd_id}")
        try:
            self.wb.remove(self.wb.active)
            ws = self.wb.create_sheet("Данные КРД")
            row = self._fill_single_krd(ws)
            self._adjust_column_widths(ws)
            self.wb.save(file_path)
            self._cleanup_temp_files()
            self._log("EXPORT", "✅ ЭКСПОРТ ОДНОЙ КРД ЗАВЕРШЁН")
            return True
        except Exception as e:
            self._log("ERROR", f"✗ ОШИБКА: {e}")
            traceback.print_exc()
            self._cleanup_temp_files()
            raise

    def _fill_single_krd(self, ws):
        row = 1
        social_data = self._load_social_data_for_krd(self.krd_id)
        row = self._fill_section(ws, "СОЦИАЛЬНО-ДЕМОГРАФИЧЕСКИЕ ДАННЫЕ", self.AVAILABLE_FIELDS["social_data"]["fields"], [social_data], row)
        if "addresses" in self.report_config.get("sections", []):
            row = self._fill_section(ws, "АДРЕСА", self.AVAILABLE_FIELDS["addresses"]["fields"], self._load_addresses_for_krd(self.krd_id), row)
        if "service_places" in self.report_config.get("sections", []):
            row = self._fill_section(ws, "МЕСТА СЛУЖБЫ", self.AVAILABLE_FIELDS["service_places"]["fields"], self._load_service_places_for_krd(self.krd_id), row)
        return row

    def _fill_section(self, ws, title, fields, data_list, start_row):
        if not data_list: return start_row
        ws.merge_cells(f"A{start_row}:Z{start_row}")
        cell = ws.cell(row=start_row, column=1, value=title)
        cell.font = self.section_font; cell.fill = self.section_fill; cell.alignment = self.header_alignment
        row = start_row + 1
        for c, (_, label) in enumerate(fields, 1):
            h_cell = ws.cell(row=row, column=c, value=label)
            h_cell.font = self.header_font; h_cell.fill = self.header_fill; h_cell.alignment = self.header_alignment; h_cell.border = self.thin_border
        row += 1
        for data in data_list:
            for c, (key, _) in enumerate(fields, 1):
                val = self._format_date(data.get(key)) if key.endswith("_date") else data.get(key, "")
                ws.cell(row=row, column=c, value=val).border = self.thin_border
            row += 1
        return row + 1